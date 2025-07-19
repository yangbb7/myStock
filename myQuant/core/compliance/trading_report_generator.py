import asyncio
import logging
from typing import Dict, List, Optional, Any, Tuple, Union
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from enum import Enum
import numpy as np
import pandas as pd
from pathlib import Path
import json
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from concurrent.futures import ThreadPoolExecutor
import warnings
warnings.filterwarnings('ignore')

class ReportType(Enum):
    TRADE_BLOTTER = "trade_blotter"
    EXECUTION_REPORT = "execution_report"
    ALLOCATION_REPORT = "allocation_report"
    COMMISSION_REPORT = "commission_report"
    BEST_EXECUTION = "best_execution"
    REGULATORY_FILING = "regulatory_filing"
    RISK_REPORT = "risk_report"
    COMPLIANCE_REPORT = "compliance_report"
    AUDIT_REPORT = "audit_report"
    EXCEPTION_REPORT = "exception_report"

class ReportFormat(Enum):
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    XML = "xml"
    JSON = "json"
    HTML = "html"
    TXT = "txt"
    REGULATORY_XML = "regulatory_xml"

class ReportFrequency(Enum):
    REAL_TIME = "real_time"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    ANNUALLY = "annually"
    ON_DEMAND = "on_demand"

class RegulatoryJurisdiction(Enum):
    SEC = "sec"           # US Securities and Exchange Commission
    FINRA = "finra"       # Financial Industry Regulatory Authority
    CFTC = "cftc"         # Commodity Futures Trading Commission
    CSRC = "csrc"         # China Securities Regulatory Commission
    FCA = "fca"           # UK Financial Conduct Authority
    ESMA = "esma"         # European Securities and Markets Authority
    MiFID = "mifid"       # Markets in Financial Instruments Directive
    EMIR = "emir"         # European Market Infrastructure Regulation
    SFTR = "sftr"         # Securities Financing Transactions Regulation
    CUSTOM = "custom"

@dataclass
class TradeRecord:
    """交易记录"""
    trade_id: str
    order_id: str
    execution_id: str
    symbol: str
    side: str  # BUY/SELL
    quantity: float
    price: float
    execution_time: datetime
    settlement_date: datetime
    counterparty: str
    venue: str
    commission: float
    fees: float
    tax: float
    net_amount: float
    currency: str
    trader_id: str
    account_id: str
    strategy_id: str
    client_id: str
    order_type: str
    time_in_force: str
    execution_quality: str
    market_impact: float
    slippage: float
    is_best_execution: bool
    regulatory_flags: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ReportConfig:
    """报告配置"""
    report_type: ReportType
    report_format: ReportFormat
    frequency: ReportFrequency
    jurisdiction: RegulatoryJurisdiction
    output_path: str
    template_path: Optional[str] = None
    include_sections: List[str] = field(default_factory=list)
    exclude_sections: List[str] = field(default_factory=list)
    date_range: Optional[Tuple[datetime, datetime]] = None
    filters: Dict[str, Any] = field(default_factory=dict)
    recipients: List[str] = field(default_factory=list)
    auto_send: bool = False
    encryption: bool = False
    digital_signature: bool = False
    retention_period: int = 2555  # days (7 years)
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ReportMetadata:
    """报告元数据"""
    report_id: str
    report_type: ReportType
    generation_time: datetime
    reporting_period: Tuple[datetime, datetime]
    total_records: int
    total_trades: int
    total_volume: float
    total_value: float
    jurisdiction: RegulatoryJurisdiction
    version: str
    generator: str
    file_hash: str
    digital_signature: Optional[str] = None
    approval_status: str = "pending"
    approved_by: Optional[str] = None
    approval_time: Optional[datetime] = None
    submission_status: str = "not_submitted"
    submission_time: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

class TradingReportGenerator:
    """
    交易报告生成系统
    
    提供全面的交易报告生成功能，支持多种监管要求、
    报告格式和自动化流程。
    """
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # 报告配置
        self.output_directory = Path(config.get('output_directory', './reports'))
        self.template_directory = Path(config.get('template_directory', './templates'))
        self.retention_period = config.get('retention_period', 2555)  # 7 years
        
        # 确保目录存在
        self.output_directory.mkdir(parents=True, exist_ok=True)
        self.template_directory.mkdir(parents=True, exist_ok=True)
        
        # 报告模板
        self.templates = {}
        self.report_schemas = {}
        
        # 生成历史
        self.generation_history = []
        self.report_registry = {}
        
        # 监管要求映射
        self.regulatory_requirements = {
            RegulatoryJurisdiction.SEC: {
                'trade_reporting': True,
                'best_execution': True,
                'order_audit_trail': True,
                'large_trader_reporting': True,
                'required_fields': ['trade_id', 'symbol', 'quantity', 'price', 'execution_time']
            },
            RegulatoryJurisdiction.FINRA: {
                'trade_reporting': True,
                'order_audit_trail': True,
                'market_maker_reporting': True,
                'required_fields': ['trade_id', 'symbol', 'side', 'quantity', 'price', 'venue']
            },
            RegulatoryJurisdiction.CSRC: {
                'trade_reporting': True,
                'position_reporting': True,
                'risk_monitoring': True,
                'required_fields': ['trade_id', 'symbol', 'quantity', 'price', 'client_id']
            },
            RegulatoryJurisdiction.MiFID: {
                'best_execution': True,
                'transaction_reporting': True,
                'order_record_keeping': True,
                'required_fields': ['trade_id', 'symbol', 'quantity', 'price', 'venue', 'counterparty']
            }
        }
        
        # 初始化模板
        self._initialize_templates()
        
    def _initialize_templates(self):
        """初始化报告模板"""
        # 基础模板配置
        self.templates = {
            ReportType.TRADE_BLOTTER: {
                'columns': [
                    'trade_id', 'execution_time', 'symbol', 'side', 'quantity',
                    'price', 'venue', 'counterparty', 'commission', 'net_amount'
                ],
                'title': 'Trade Blotter Report',
                'description': 'Comprehensive trade execution report'
            },
            ReportType.EXECUTION_REPORT: {
                'columns': [
                    'order_id', 'execution_id', 'symbol', 'side', 'quantity',
                    'price', 'execution_time', 'venue', 'execution_quality'
                ],
                'title': 'Execution Quality Report',
                'description': 'Analysis of execution quality and best execution'
            },
            ReportType.BEST_EXECUTION: {
                'columns': [
                    'symbol', 'venue', 'avg_price', 'volume', 'market_share',
                    'execution_quality', 'cost_analysis', 'price_improvement'
                ],
                'title': 'Best Execution Report',
                'description': 'Best execution analysis and venue comparison'
            },
            ReportType.REGULATORY_FILING: {
                'columns': [
                    'report_id', 'filing_type', 'jurisdiction', 'filing_date',
                    'status', 'submission_method', 'confirmation_number'
                ],
                'title': 'Regulatory Filing Report',
                'description': 'Regulatory submission tracking and status'
            }
        }
    
    async def generate_report(self, 
                            config: ReportConfig,
                            data: List[TradeRecord],
                            metadata: Optional[Dict[str, Any]] = None) -> str:
        """生成交易报告"""
        try:
            self.logger.info(f"开始生成报告: {config.report_type.value}")
            
            # 数据验证和预处理
            validated_data = await self._validate_data(data, config)
            
            # 应用过滤器
            filtered_data = await self._apply_filters(validated_data, config)
            
            # 生成报告元数据
            report_metadata = await self._generate_metadata(filtered_data, config)
            
            # 根据格式生成报告
            if config.report_format == ReportFormat.PDF:
                file_path = await self._generate_pdf_report(filtered_data, config, report_metadata)
            elif config.report_format == ReportFormat.EXCEL:
                file_path = await self._generate_excel_report(filtered_data, config, report_metadata)
            elif config.report_format == ReportFormat.CSV:
                file_path = await self._generate_csv_report(filtered_data, config, report_metadata)
            elif config.report_format == ReportFormat.XML:
                file_path = await self._generate_xml_report(filtered_data, config, report_metadata)
            elif config.report_format == ReportFormat.JSON:
                file_path = await self._generate_json_report(filtered_data, config, report_metadata)
            elif config.report_format == ReportFormat.REGULATORY_XML:
                file_path = await self._generate_regulatory_xml(filtered_data, config, report_metadata)
            else:
                raise ValueError(f"不支持的报告格式: {config.report_format}")
            
            # 注册报告
            await self._register_report(file_path, report_metadata)
            
            # 自动发送
            if config.auto_send and config.recipients:
                await self._send_report(file_path, config)
            
            self.logger.info(f"报告生成完成: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.error(f"报告生成失败: {e}")
            raise
    
    async def _validate_data(self, data: List[TradeRecord], config: ReportConfig) -> List[TradeRecord]:
        """验证数据"""
        validated_data = []
        
        # 获取监管要求
        regulatory_req = self.regulatory_requirements.get(
            config.jurisdiction, {}
        )
        required_fields = regulatory_req.get('required_fields', [])
        
        for record in data:
            # 检查必需字段
            missing_fields = []
            for field in required_fields:
                if not hasattr(record, field) or getattr(record, field) is None:
                    missing_fields.append(field)
            
            if missing_fields:
                self.logger.warning(f"交易记录 {record.trade_id} 缺少必需字段: {missing_fields}")
                continue
            
            # 数据类型验证
            if not isinstance(record.quantity, (int, float)) or record.quantity <= 0:
                self.logger.warning(f"交易记录 {record.trade_id} 数量无效: {record.quantity}")
                continue
            
            if not isinstance(record.price, (int, float)) or record.price <= 0:
                self.logger.warning(f"交易记录 {record.trade_id} 价格无效: {record.price}")
                continue
            
            # 日期验证
            if not isinstance(record.execution_time, datetime):
                self.logger.warning(f"交易记录 {record.trade_id} 执行时间无效")
                continue
            
            validated_data.append(record)
        
        self.logger.info(f"数据验证完成: {len(validated_data)}/{len(data)} 记录通过验证")
        return validated_data
    
    async def _apply_filters(self, data: List[TradeRecord], config: ReportConfig) -> List[TradeRecord]:
        """应用过滤器"""
        filtered_data = data
        
        # 日期范围过滤
        if config.date_range:
            start_date, end_date = config.date_range
            filtered_data = [
                record for record in filtered_data
                if start_date <= record.execution_time <= end_date
            ]
        
        # 其他过滤器
        for filter_key, filter_value in config.filters.items():
            if filter_key == 'symbol':
                filtered_data = [
                    record for record in filtered_data
                    if record.symbol in filter_value
                ]
            elif filter_key == 'venue':
                filtered_data = [
                    record for record in filtered_data
                    if record.venue in filter_value
                ]
            elif filter_key == 'trader_id':
                filtered_data = [
                    record for record in filtered_data
                    if record.trader_id in filter_value
                ]
            elif filter_key == 'min_quantity':
                filtered_data = [
                    record for record in filtered_data
                    if record.quantity >= filter_value
                ]
            elif filter_key == 'min_value':
                filtered_data = [
                    record for record in filtered_data
                    if record.quantity * record.price >= filter_value
                ]
        
        self.logger.info(f"过滤器应用完成: {len(filtered_data)} 记录")
        return filtered_data
    
    async def _generate_metadata(self, data: List[TradeRecord], config: ReportConfig) -> ReportMetadata:
        """生成报告元数据"""
        # 计算统计信息
        total_records = len(data)
        total_trades = total_records
        total_volume = sum(record.quantity for record in data)
        total_value = sum(record.quantity * record.price for record in data)
        
        # 确定报告期间
        if data:
            start_date = min(record.execution_time for record in data)
            end_date = max(record.execution_time for record in data)
        else:
            start_date = end_date = datetime.now()
        
        # 生成报告ID
        report_id = f"{config.report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        return ReportMetadata(
            report_id=report_id,
            report_type=config.report_type,
            generation_time=datetime.now(),
            reporting_period=(start_date, end_date),
            total_records=total_records,
            total_trades=total_trades,
            total_volume=total_volume,
            total_value=total_value,
            jurisdiction=config.jurisdiction,
            version="1.0",
            generator="myStock Trading System",
            file_hash="",  # 将在文件生成后计算
            approval_status="pending",
            submission_status="not_submitted"
        )
    
    async def _generate_pdf_report(self, data: List[TradeRecord], 
                                 config: ReportConfig, 
                                 metadata: ReportMetadata) -> str:
        """生成PDF报告"""
        file_path = self.output_directory / f"{metadata.report_id}.pdf"
        
        # 创建PDF文档
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # 居中
        )
        
        template = self.templates.get(config.report_type, {})
        title = template.get('title', 'Trading Report')
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # 报告信息
        info_data = [
            ['Report ID:', metadata.report_id],
            ['Generation Time:', metadata.generation_time.strftime('%Y-%m-%d %H:%M:%S')],
            ['Reporting Period:', f"{metadata.reporting_period[0].strftime('%Y-%m-%d')} to {metadata.reporting_period[1].strftime('%Y-%m-%d')}"],
            ['Total Records:', str(metadata.total_records)],
            ['Total Volume:', f"{metadata.total_volume:,.2f}"],
            ['Total Value:', f"${metadata.total_value:,.2f}"],
            ['Jurisdiction:', metadata.jurisdiction.value.upper()]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 30))
        
        # 数据表格
        if data:
            # 表头
            columns = template.get('columns', ['trade_id', 'symbol', 'quantity', 'price'])
            table_data = [columns]
            
            # 数据行
            for record in data:
                row = []
                for column in columns:
                    value = getattr(record, column, '')
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, float):
                        value = f"{value:,.2f}"
                    row.append(str(value))
                table_data.append(row)
            
            # 创建表格
            data_table = Table(table_data)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(data_table)
        
        # 生成PDF
        doc.build(story)
        
        return str(file_path)
    
    async def _generate_excel_report(self, data: List[TradeRecord], 
                                   config: ReportConfig, 
                                   metadata: ReportMetadata) -> str:
        """生成Excel报告"""
        file_path = self.output_directory / f"{metadata.report_id}.xlsx"
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading Report"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 写入元数据
        ws['A1'] = 'Report Information'
        ws['A1'].font = Font(bold=True, size=14)
        
        info_data = [
            ['Report ID:', metadata.report_id],
            ['Generation Time:', metadata.generation_time.strftime('%Y-%m-%d %H:%M:%S')],
            ['Reporting Period:', f"{metadata.reporting_period[0].strftime('%Y-%m-%d')} to {metadata.reporting_period[1].strftime('%Y-%m-%d')}"],
            ['Total Records:', metadata.total_records],
            ['Total Volume:', metadata.total_volume],
            ['Total Value:', metadata.total_value],
            ['Jurisdiction:', metadata.jurisdiction.value.upper()]
        ]
        
        for i, (key, value) in enumerate(info_data, start=2):
            ws.cell(row=i, column=1, value=key).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # 数据表格
        if data:
            # 表头
            template = self.templates.get(config.report_type, {})
            columns = template.get('columns', ['trade_id', 'symbol', 'quantity', 'price'])
            
            start_row = len(info_data) + 4
            
            # 写入列头
            for col_num, column in enumerate(columns, start=1):
                cell = ws.cell(row=start_row, column=col_num, value=column.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
            
            # 写入数据
            for row_num, record in enumerate(data, start=start_row+1):
                for col_num, column in enumerate(columns, start=1):
                    value = getattr(record, column, '')
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(file_path)
        
        return str(file_path)
    
    async def _generate_csv_report(self, data: List[TradeRecord], 
                                 config: ReportConfig, 
                                 metadata: ReportMetadata) -> str:
        """生成CSV报告"""
        file_path = self.output_directory / f"{metadata.report_id}.csv"
        
        if not data:
            # 创建空文件
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                f.write("# No data available\n")
            return str(file_path)
        
        # 准备数据
        template = self.templates.get(config.report_type, {})
        columns = template.get('columns', ['trade_id', 'symbol', 'quantity', 'price'])
        
        # 创建DataFrame
        data_dict = {}
        for column in columns:
            data_dict[column] = []
            for record in data:
                value = getattr(record, column, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                data_dict[column].append(value)
        
        df = pd.DataFrame(data_dict)
        
        # 保存CSV
        df.to_csv(file_path, index=False, encoding='utf-8')
        
        return str(file_path)
    
    async def _generate_xml_report(self, data: List[TradeRecord], 
                                 config: ReportConfig, 
                                 metadata: ReportMetadata) -> str:
        """生成XML报告"""
        file_path = self.output_directory / f"{metadata.report_id}.xml"
        
        # 创建根元素
        root = ET.Element("TradingReport")
        
        # 添加元数据
        metadata_elem = ET.SubElement(root, "Metadata")
        ET.SubElement(metadata_elem, "ReportID").text = metadata.report_id
        ET.SubElement(metadata_elem, "GenerationTime").text = metadata.generation_time.isoformat()
        ET.SubElement(metadata_elem, "ReportingPeriod").text = f"{metadata.reporting_period[0].isoformat()}/{metadata.reporting_period[1].isoformat()}"
        ET.SubElement(metadata_elem, "TotalRecords").text = str(metadata.total_records)
        ET.SubElement(metadata_elem, "TotalVolume").text = str(metadata.total_volume)
        ET.SubElement(metadata_elem, "TotalValue").text = str(metadata.total_value)
        ET.SubElement(metadata_elem, "Jurisdiction").text = metadata.jurisdiction.value
        
        # 添加数据
        trades_elem = ET.SubElement(root, "Trades")
        
        for record in data:
            trade_elem = ET.SubElement(trades_elem, "Trade")
            
            # 添加交易字段
            for field_name in ['trade_id', 'symbol', 'side', 'quantity', 'price', 'execution_time', 'venue']:
                if hasattr(record, field_name):
                    value = getattr(record, field_name)
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, float):
                        value = f"{value:.6f}"
                    ET.SubElement(trade_elem, field_name.replace('_', '')).text = str(value)
        
        # 保存XML
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        return str(file_path)
    
    async def _generate_json_report(self, data: List[TradeRecord], 
                                  config: ReportConfig, 
                                  metadata: ReportMetadata) -> str:
        """生成JSON报告"""
        file_path = self.output_directory / f"{metadata.report_id}.json"
        
        # 准备数据
        report_data = {
            "metadata": {
                "report_id": metadata.report_id,
                "report_type": config.report_type.value,
                "generation_time": metadata.generation_time.isoformat(),
                "reporting_period": {
                    "start": metadata.reporting_period[0].isoformat(),
                    "end": metadata.reporting_period[1].isoformat()
                },
                "total_records": metadata.total_records,
                "total_volume": metadata.total_volume,
                "total_value": metadata.total_value,
                "jurisdiction": metadata.jurisdiction.value
            },
            "trades": []
        }
        
        # 添加交易数据
        for record in data:
            trade_data = {
                "trade_id": record.trade_id,
                "symbol": record.symbol,
                "side": record.side,
                "quantity": record.quantity,
                "price": record.price,
                "execution_time": record.execution_time.isoformat(),
                "venue": record.venue,
                "commission": record.commission,
                "net_amount": record.net_amount,
                "currency": record.currency
            }
            report_data["trades"].append(trade_data)
        
        # 保存JSON
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        return str(file_path)
    
    async def _generate_regulatory_xml(self, data: List[TradeRecord], 
                                     config: ReportConfig, 
                                     metadata: ReportMetadata) -> str:
        """生成监管XML报告"""
        file_path = self.output_directory / f"{metadata.report_id}_regulatory.xml"
        
        # 根据监管要求生成XML
        if config.jurisdiction == RegulatoryJurisdiction.MiFID:
            return await self._generate_mifid_xml(data, config, metadata, file_path)
        elif config.jurisdiction == RegulatoryJurisdiction.EMIR:
            return await self._generate_emir_xml(data, config, metadata, file_path)
        elif config.jurisdiction == RegulatoryJurisdiction.SEC:
            return await self._generate_sec_xml(data, config, metadata, file_path)
        else:
            # 默认格式
            return await self._generate_xml_report(data, config, metadata)
    
    async def _generate_mifid_xml(self, data: List[TradeRecord], 
                                config: ReportConfig, 
                                metadata: ReportMetadata, 
                                file_path: Path) -> str:
        """生成MiFID II格式XML"""
        # MiFID II特定的XML结构
        root = ET.Element("Document", xmlns="urn:iso:std:iso:20022:tech:xsd:auth.036.001.02")
        
        # 消息头
        fin_instrm_rpt_tx_rpt = ET.SubElement(root, "FinInstrmRptgTxRpt")
        grp_hdr = ET.SubElement(fin_instrm_rpt_tx_rpt, "GrpHdr")
        ET.SubElement(grp_hdr, "MsgId").text = metadata.report_id
        ET.SubElement(grp_hdr, "CreDtTm").text = metadata.generation_time.isoformat()
        ET.SubElement(grp_hdr, "NbOfTxs").text = str(metadata.total_records)
        
        # 交易报告
        for record in data:
            tx_rpt = ET.SubElement(fin_instrm_rpt_tx_rpt, "TxRpt")
            
            # 交易标识
            ET.SubElement(tx_rpt, "TxId").text = record.trade_id
            ET.SubElement(tx_rpt, "ExctgPty").text = record.counterparty
            ET.SubElement(tx_rpt, "SubmitgPty").text = "TRADING_SYSTEM"
            
            # 金融工具
            fin_instrm = ET.SubElement(tx_rpt, "FinInstrm")
            ET.SubElement(fin_instrm, "Id").text = record.symbol
            
            # 交易详情
            ET.SubElement(tx_rpt, "TradgDt").text = record.execution_time.date().isoformat()
            ET.SubElement(tx_rpt, "TradgTm").text = record.execution_time.time().isoformat()
            ET.SubElement(tx_rpt, "TradgCpcty").text = "AOTC"  # Acting on own account
            ET.SubElement(tx_rpt, "Qty").text = str(record.quantity)
            ET.SubElement(tx_rpt, "Pric").text = str(record.price)
            ET.SubElement(tx_rpt, "TradVn").text = record.venue
        
        # 保存文件
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        return str(file_path)
    
    async def _generate_emir_xml(self, data: List[TradeRecord], 
                               config: ReportConfig, 
                               metadata: ReportMetadata, 
                               file_path: Path) -> str:
        """生成EMIR格式XML"""
        # EMIR特定的XML结构（简化版）
        root = ET.Element("Document", xmlns="urn:iso:std:iso:20022:tech:xsd:auth.018.001.02")
        
        deriv_rpt = ET.SubElement(root, "DerivRpt")
        rpt_hdr = ET.SubElement(deriv_rpt, "RptHdr")
        ET.SubElement(rpt_hdr, "RptId").text = metadata.report_id
        ET.SubElement(rpt_hdr, "RptgDt").text = metadata.generation_time.date().isoformat()
        
        # 交易报告（仅衍生品）
        for record in data:
            if 'derivative' in record.metadata.get('asset_type', '').lower():
                tx_dtls = ET.SubElement(deriv_rpt, "TxDtls")
                ET.SubElement(tx_dtls, "TxId").text = record.trade_id
                ET.SubElement(tx_dtls, "ExctnDtTm").text = record.execution_time.isoformat()
                ET.SubElement(tx_dtls, "NtnlAmt").text = str(record.quantity * record.price)
        
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        return str(file_path)
    
    async def _generate_sec_xml(self, data: List[TradeRecord], 
                              config: ReportConfig, 
                              metadata: ReportMetadata, 
                              file_path: Path) -> str:
        """生成SEC格式XML"""
        # SEC特定的XML结构（简化版）
        root = ET.Element("SecuritiesTransactionReport")
        
        # 报告头
        header = ET.SubElement(root, "ReportHeader")
        ET.SubElement(header, "ReportId").text = metadata.report_id
        ET.SubElement(header, "ReportingFirm").text = "TRADING_SYSTEM"
        ET.SubElement(header, "ReportingDate").text = metadata.generation_time.date().isoformat()
        ET.SubElement(header, "TotalTransactions").text = str(metadata.total_records)
        
        # 交易数据
        transactions = ET.SubElement(root, "Transactions")
        
        for record in data:
            transaction = ET.SubElement(transactions, "Transaction")
            ET.SubElement(transaction, "TransactionId").text = record.trade_id
            ET.SubElement(transaction, "Symbol").text = record.symbol
            ET.SubElement(transaction, "Side").text = record.side
            ET.SubElement(transaction, "Quantity").text = str(record.quantity)
            ET.SubElement(transaction, "Price").text = str(record.price)
            ET.SubElement(transaction, "ExecutionTime").text = record.execution_time.isoformat()
            ET.SubElement(transaction, "Venue").text = record.venue
            ET.SubElement(transaction, "Commission").text = str(record.commission)
        
        tree = ET.ElementTree(root)
        tree.write(file_path, encoding='utf-8', xml_declaration=True)
        
        return str(file_path)
    
    async def _register_report(self, file_path: str, metadata: ReportMetadata):
        """注册报告"""
        # 计算文件哈希
        import hashlib
        with open(file_path, 'rb') as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
        
        metadata.file_hash = file_hash
        
        # 注册到报告注册表
        self.report_registry[metadata.report_id] = {
            'metadata': metadata,
            'file_path': file_path,
            'registration_time': datetime.now(),
            'status': 'generated'
        }
        
        # 添加到生成历史
        self.generation_history.append({
            'report_id': metadata.report_id,
            'generation_time': metadata.generation_time,
            'file_path': file_path,
            'status': 'success'
        })
        
        self.logger.info(f"报告已注册: {metadata.report_id}")
    
    async def _send_report(self, file_path: str, config: ReportConfig):
        """发送报告"""
        # 简化实现 - 实际应该集成邮件系统
        self.logger.info(f"报告发送给: {config.recipients}")
        
        # 这里可以集成SMTP、SFTP、API等发送方式
        for recipient in config.recipients:
            self.logger.info(f"发送报告到: {recipient}")
    
    async def get_report_status(self, report_id: str) -> Dict[str, Any]:
        """获取报告状态"""
        if report_id not in self.report_registry:
            return {'status': 'not_found'}
        
        report_info = self.report_registry[report_id]
        return {
            'report_id': report_id,
            'status': report_info['status'],
            'file_path': report_info['file_path'],
            'generation_time': report_info['metadata'].generation_time,
            'approval_status': report_info['metadata'].approval_status,
            'submission_status': report_info['metadata'].submission_status
        }
    
    async def approve_report(self, report_id: str, approver: str) -> bool:
        """批准报告"""
        if report_id not in self.report_registry:
            return False
        
        report_info = self.report_registry[report_id]
        report_info['metadata'].approval_status = 'approved'
        report_info['metadata'].approved_by = approver
        report_info['metadata'].approval_time = datetime.now()
        report_info['status'] = 'approved'
        
        self.logger.info(f"报告已批准: {report_id} by {approver}")
        return True
    
    async def submit_report(self, report_id: str, submission_method: str = 'manual') -> bool:
        """提交报告"""
        if report_id not in self.report_registry:
            return False
        
        report_info = self.report_registry[report_id]
        
        if report_info['metadata'].approval_status != 'approved':
            self.logger.error(f"报告未批准，无法提交: {report_id}")
            return False
        
        report_info['metadata'].submission_status = 'submitted'
        report_info['metadata'].submission_time = datetime.now()
        report_info['status'] = 'submitted'
        
        self.logger.info(f"报告已提交: {report_id} via {submission_method}")
        return True
    
    async def generate_summary_report(self, period: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """生成汇总报告"""
        start_date, end_date = period
        
        # 筛选期间内的报告
        period_reports = [
            report for report in self.report_registry.values()
            if start_date <= report['metadata'].generation_time <= end_date
        ]
        
        # 统计信息
        total_reports = len(period_reports)
        approved_reports = len([r for r in period_reports if r['metadata'].approval_status == 'approved'])
        submitted_reports = len([r for r in period_reports if r['metadata'].submission_status == 'submitted'])
        
        # 按类型分组
        report_by_type = {}
        for report in period_reports:
            report_type = report['metadata'].report_type.value
            if report_type not in report_by_type:
                report_by_type[report_type] = 0
            report_by_type[report_type] += 1
        
        # 按监管机构分组
        report_by_jurisdiction = {}
        for report in period_reports:
            jurisdiction = report['metadata'].jurisdiction.value
            if jurisdiction not in report_by_jurisdiction:
                report_by_jurisdiction[jurisdiction] = 0
            report_by_jurisdiction[jurisdiction] += 1
        
        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'summary': {
                'total_reports': total_reports,
                'approved_reports': approved_reports,
                'submitted_reports': submitted_reports,
                'pending_approval': total_reports - approved_reports,
                'pending_submission': approved_reports - submitted_reports
            },
            'breakdown': {
                'by_type': report_by_type,
                'by_jurisdiction': report_by_jurisdiction
            },
            'reports': [
                {
                    'report_id': report['metadata'].report_id,
                    'type': report['metadata'].report_type.value,
                    'generation_time': report['metadata'].generation_time.isoformat(),
                    'approval_status': report['metadata'].approval_status,
                    'submission_status': report['metadata'].submission_status
                }
                for report in period_reports
            ]
        }
    
    async def cleanup_old_reports(self, retention_days: int = None):
        """清理过期报告"""
        if retention_days is None:
            retention_days = self.retention_period
        
        cutoff_date = datetime.now() - timedelta(days=retention_days)
        
        # 查找过期报告
        expired_reports = [
            report_id for report_id, report_info in self.report_registry.items()
            if report_info['metadata'].generation_time < cutoff_date
        ]
        
        # 删除过期报告
        for report_id in expired_reports:
            try:
                report_info = self.report_registry[report_id]
                file_path = Path(report_info['file_path'])
                
                if file_path.exists():
                    file_path.unlink()
                
                del self.report_registry[report_id]
                self.logger.info(f"已删除过期报告: {report_id}")
                
            except Exception as e:
                self.logger.error(f"删除过期报告失败 {report_id}: {e}")
        
        return len(expired_reports)
    
    async def validate_regulatory_compliance(self, report_id: str) -> Dict[str, Any]:
        """验证监管合规性"""
        if report_id not in self.report_registry:
            return {'valid': False, 'error': 'Report not found'}
        
        report_info = self.report_registry[report_id]
        metadata = report_info['metadata']
        
        # 获取监管要求
        jurisdiction = metadata.jurisdiction
        requirements = self.regulatory_requirements.get(jurisdiction, {})
        
        compliance_issues = []
        
        # 检查必需字段
        required_fields = requirements.get('required_fields', [])
        # 这里应该检查报告数据是否包含所有必需字段
        
        # 检查时效性
        if metadata.generation_time < datetime.now() - timedelta(days=1):
            compliance_issues.append("Report may be outdated")
        
        # 检查批准状态
        if metadata.approval_status != 'approved':
            compliance_issues.append("Report not approved")
        
        # 检查数字签名
        if requirements.get('digital_signature_required', False) and not metadata.digital_signature:
            compliance_issues.append("Digital signature required")
        
        return {
            'valid': len(compliance_issues) == 0,
            'compliance_issues': compliance_issues,
            'jurisdiction': jurisdiction.value,
            'requirements_met': len(required_fields) - len(compliance_issues),
            'total_requirements': len(required_fields)
        }